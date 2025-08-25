from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, NamedStyle, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import win32com.client as win32
from openpyxl.utils import get_column_letter
from datetime import datetime
import pandas as pd
import time
import os
import glob
import random
import xlwings as xw
import inspect
import sys

# Obtener la ruta base del directorio donde est√° el script
base_dir = os.path.dirname(os.path.abspath(__file__))

# Definir rutas a las carpetas y archivos
input_folder_excel = os.path.join(base_dir, "data", "input", "Deudas")
output_folder_csv = os.path.join(base_dir, "data", "input", "DeudasCSV")
output_file_csv = os.path.join(base_dir, "data", "Resumen_deudas.csv")
output_file_xlsx = os.path.join(base_dir, "data", "Resumen_deudas.xlsx")

output_folder_pdf = os.path.join(base_dir, "data", "Reportes")
imagen = os.path.join(base_dir, "data", "imagen.png")

# Leer el archivo Excel
input_excel_clientes = os.path.join(base_dir, "data", "input", "clientes.xlsx")
df = pd.read_excel(input_excel_clientes)

# Suposici√≥n de nombres de columnas
cuit_login_list = df['CUIT para ingresar'].tolist()
cuit_represent_list = df['CUIT representado'].tolist()
password_list = df['Contrase√±a'].tolist()
download_list = df['Ubicacion descarga'].tolist()
clientes_list = df['Cliente'].tolist()

driver = None

def configurar_nuevo_navegador():
    global driver
    
    # Configuraci√≥n de opciones de Chrome
    options = Options()
    options.add_argument("--start-maximized")
    
    # Configurar preferencias de descarga
    prefs = {
        "download.prompt_for_download": True,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    options.add_experimental_option("prefs", prefs)
    
    # Inicializar driver nuevo
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    
    print("‚úÖ Nuevo navegador Chrome configurado")
    return driver

def cerrar_sesion_y_navegador():
    global driver
    
    try:
        print("\n--- INICIANDO CIERRE COMPLETO DE SESI√ìN ---")
        
        # PASO 1: Verificar cu√°ntas pesta√±as est√°n abiertas
        window_handles = driver.window_handles
        num_pestanas = len(window_handles)
        print(f"üìä Pesta√±as abiertas detectadas: {num_pestanas}")
        
        # PASO 2: Si hay m√°s de 1 pesta√±a, cerrar las adicionales
        if num_pestanas > 1:
            print(f"üîÑ Cerrando {num_pestanas - 1} pesta√±as adicionales...")
            
            # Ir a la √∫ltima pesta√±a (SCT) y cerrarla
            for i in range(num_pestanas - 1, 0, -1):  # Desde la √∫ltima hacia la segunda
                try:
                    driver.switch_to.window(window_handles[i])
                    print(f"üóÇÔ∏è Cerrando pesta√±a {i + 1}: {driver.title[:50]}...")
                    driver.close()
                    time.sleep(2)
                except Exception as e:
                    print(f"‚ö†Ô∏è Error cerrando pesta√±a {i + 1}: {e}")
            
            # Volver a la pesta√±a principal (√≠ndice 0)
            driver.switch_to.window(window_handles[0])
            print("‚úÖ Vuelto a la pesta√±a principal")
            time.sleep(2)
        
        # PASO 3: Intentar cerrar sesi√≥n en AFIP desde la pesta√±a principal
        try:
            print("üîí Intentando cerrar sesi√≥n en AFIP...")
            
            # Buscar el icono de contribuyente AFIP
            icono_contribuyente = driver.find_element(By.ID, "iconoChicoContribuyenteAFIP")
            icono_contribuyente.click()
            time.sleep(2)
            
            # Buscar y hacer clic en el bot√≥n de salir
            boton_salir = driver.find_element(By.XPATH, '//*[@id="contBtnContribuyente"]/div[6]/button/div/div[2]')
            boton_salir.click()
            time.sleep(2)
            
            print("‚úÖ Sesi√≥n cerrada exitosamente en AFIP")
            
        except Exception as e:
            print(f"‚ö†Ô∏è No se pudo cerrar sesi√≥n en AFIP (puede que no est√© logueado): {e}")
        
        # PASO 4: Cerrar el navegador completamente
        print("üåê Cerrando navegador completamente...")
        driver.quit()
        driver = None
        print("‚úÖ Navegador cerrado exitosamente")
        
    except Exception as e:
        print(f"üö® Error durante cierre completo: {e}")
        # Forzar cierre del navegador en caso de error
        try:
            if driver:
                driver.quit()
                driver = None
        except:
            pass
    
    print("--- CIERRE COMPLETO FINALIZADO ---\n")

# Crear el archivo de resultados
resultados = []

def human_typing(element, text):
    for char in str(text):
        element.send_keys(char)
        time.sleep(random.uniform(0.01, 0.05))

def actualizar_excel(row_index, mensaje):
    """Actualiza la √∫ltima columna del archivo Excel con un mensaje de error."""
    df.at[row_index, 'Error'] = mensaje
    df.to_excel(input_excel_clientes, index=False)

def verificar_funciones_disponibles():
    """Verifica que todas las funciones necesarias est√©n disponibles."""
    funciones_necesarias = ['procesar_excel', 'aplicar_filtros_deudas', 'generar_pdf_desde_dataframe']
    
    current_module = sys.modules[__name__]
    
    print("=== VERIFICACI√ìN DE FUNCIONES ===")
    for func_name in funciones_necesarias:
        if hasattr(current_module, func_name):
            print(f"‚úì Funci√≥n {func_name} disponible")
        else:
            print(f"‚úó Funci√≥n {func_name} NO disponible")
    
    # Mostrar algunas funciones disponibles
    all_functions = [name for name, obj in inspect.getmembers(current_module) if inspect.isfunction(obj)]
    print(f"Total funciones disponibles: {len(all_functions)}")

def aplicar_filtros_deudas(df, cliente):
    """Aplica los mismos filtros que se aplicaban al Excel."""
    try:
        print(f"\n--- APLICANDO FILTROS PARA {cliente} ---")
        print(f"Datos originales: {len(df)} registros")
        
        # Lista de impuestos a incluir (misma que se usa en procesar_excel)
        impuestos_incluir = [
            'ganancias sociedades',
            'iva',
            'bp-acciones o participaciones',
            'sicore-impto.a las ganancias',
            'empleador-aportes seg. social',
            'contribuciones seg. social',
            'ret art 79 ley gcias inc a,byc',
            'renatea',
            'ganancias personas fisicas',
            'ganancia minima presunta',
            'seguro de vida colectivo',
            'regimenes de informacion',
            'imp s/deb y cred en cta cte',
            'presentac. dj ret. y/o percep',
            'contrib.vales aliment.l.24700',
            'aportes obra social',
            'aseg.riesgo de trabajo l 24557',
            'contribuciones obra social',
            'contribuciones renatea',
            'derecho exportacion servicios',
            'impto.s/bienes personales',
            'multas infracciones formales',
            'retenciones contrib.seg.social',
            'sicore - retenciones y percepc'
        ]
        
        # Filtrar por impuestos (si existe la columna)
        if 'Impuesto' in df.columns:
            condicion_impuestos = df['Impuesto'].str.contains('|'.join(impuestos_incluir), case=False, na=False)
            df_filtrado = df[condicion_impuestos].copy()
            print(f"Despu√©s de filtrar impuestos: {len(df_filtrado)} registros")
        else:
            df_filtrado = df.copy()
            print("No se encontr√≥ columna 'Impuesto', manteniendo todos los registros")
        
        # Filtrar por fechas (si existe la columna de vencimiento)
        columnas_fecha = ['Vencimiento', 'Fecha Vencimiento', 'FechaVencimiento', 'Fecha de Vencimiento']
        columna_fecha_encontrada = None
        
        for col in columnas_fecha:
            if col in df_filtrado.columns:
                columna_fecha_encontrada = col
                break
        
        if columna_fecha_encontrada:
            print(f"Aplicando filtro de fechas en columna: {columna_fecha_encontrada}")
            
            # Obtener fechas
            fecha_actual = pd.Timestamp.now().date()
            a√±o_actual = fecha_actual.year
            fecha_inicio = pd.Timestamp(year=a√±o_actual - 7, month=1, day=1).date()
            
            # Procesar fechas
            df_filtrado['fecha_procesada'] = pd.to_datetime(
                df_filtrado[columna_fecha_encontrada], 
                errors='coerce',
                dayfirst=True
            ).dt.date
            
            # Aplicar filtro de fechas
            mascara_fechas = (
                df_filtrado['fecha_procesada'].notna() & 
                (df_filtrado['fecha_procesada'] >= fecha_inicio) &
                (df_filtrado['fecha_procesada'] <= fecha_actual)
            )
            
            df_filtrado = df_filtrado[mascara_fechas]
            df_filtrado = df_filtrado.drop(['fecha_procesada'], axis=1)
            
            print(f"Despu√©s de filtrar fechas: {len(df_filtrado)} registros")
        
        # Eliminar columnas innecesarias
        columnas_eliminar = ['Int. punitorio', 'Concepto', 'Subconcepto', 'Establecimiento']
        for col in columnas_eliminar:
            if col in df_filtrado.columns:
                df_filtrado = df_filtrado.drop(col, axis=1)
                print(f"Columna '{col}' eliminada")
        
        print(f"Datos finales: {len(df_filtrado)} registros")
        return df_filtrado
        
    except Exception as e:
        print(f"Error aplicando filtros: {e}")
        return df

def crear_pdf_simple(excel_file, output_pdf, cliente):
    """Funci√≥n alternativa para crear PDF si procesar_excel no est√° disponible."""
    try:
        print("Usando m√©todo alternativo para generar PDF...")
        
        # Leer el Excel
        df = pd.read_excel(excel_file)
        
        # Crear un mensaje simple
        if len(df) == 0:
            mensaje = f"Reporte de Deudas SCT - {cliente}\n\nNo se encontraron deudas."
        else:
            mensaje = f"Reporte de Deudas SCT - {cliente}\n\n{len(df)} registros encontrados."
        
        # Crear un archivo de texto temporal
        txt_file = output_pdf.replace('.pdf', '_temp.txt')
        with open(txt_file, 'w', encoding='utf-8') as f:
            f.write(mensaje)
        
        print(f"PDF alternativo creado: {output_pdf}")
        
        # Limpiar archivo temporal
        try:
            os.remove(txt_file)
        except:
            pass
            
    except Exception as e:
        print(f"Error en m√©todo alternativo: {e}")

def generar_pdf_desde_dataframe(df, cliente, ruta_pdf):
    """Genera PDF directamente desde DataFrame - versi√≥n corregida."""
    try:
        print(f"\n--- GENERANDO PDF PARA {cliente} ---")
        
        # Crear Excel temporal para usar la funci√≥n existente
        temp_excel = ruta_pdf.replace('.pdf', '.xlsx').replace('Reporte - ', 'temp_excel_')
        
        if len(df) > 0:
            df.to_excel(temp_excel, index=False)
            print(f"DataFrame con {len(df)} registros guardado en Excel temporal")
        else:
            # Crear Excel vac√≠o con estructura b√°sica
            df_vacio = pd.DataFrame(columns=['Impuesto', 'Per√≠odo', 'Ant/Cuota', 'Vencimiento', 'Saldo', 'Int. Resarcitorios'])
            df_vacio.to_excel(temp_excel, index=False)
            print("Excel vac√≠o creado para PDF vac√≠o")
        
        # VERIFICAR si procesar_excel est√° disponible
        current_module = sys.modules[__name__]
        if hasattr(current_module, 'procesar_excel'):
            print("‚úì Funci√≥n procesar_excel encontrada")
            imagen_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "imagen.png")
            procesar_excel(temp_excel, ruta_pdf, imagen_path)
        else:
            print("‚úó Funci√≥n procesar_excel NO encontrada")
            # Como alternativa, crear un PDF simple
            crear_pdf_simple(temp_excel, ruta_pdf, cliente)
        
        # Limpiar archivo temporal
        try:
            os.remove(temp_excel)
            print("Archivo temporal eliminado")
        except:
            pass
            
        print(f"‚úì PDF generado exitosamente: {ruta_pdf}")
        
    except Exception as e:
        print(f"Error generando PDF: {e}")
        import traceback
        traceback.print_exc()

def formatear_numero_argentino(valor):
    # Convierte un n√∫mero a formato argentino: 1.234.567,89
    try:
        # Convertir a float
        num = float(valor) if valor != '' and valor is not None else 0.0
        
        # Formatear con 2 decimales
        if num == 0:
            return "0,00"
        
        # Usar formato argentino: separador de miles punto, decimal coma
        if num < 0:
            return f"-{abs(num):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        else:
            return f"{num:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
    except (ValueError, TypeError):
        return "0,00"

def convertir_argentino_a_float(valor_str):
    # Convierte formato argentino a float: 1.234,56 ‚Üí 1234.56
    try:
        if not valor_str or valor_str in ['', '-', 'N/A', '0,00']:
            return 0.0
        
        # Limpiar y convertir formato argentino a float
        valor_limpio = str(valor_str).replace('.', '').replace(',', '.')
        return float(valor_limpio)
    except (ValueError, TypeError):
        return 0.0

def procesar_excel(excel_file, output_pdf, imagen):
    try:
        nombre_archivo = os.path.basename(excel_file)

        # Limpiar el nombre temporal para extraer el cliente correctamente
        if nombre_archivo.startswith("temp_excel_"):
            cliente = nombre_archivo.replace("temp_excel_", "").replace(".xlsx", "")
        elif " - " in nombre_archivo:
            cliente = nombre_archivo.split(" - ")[1].replace(".xlsx", "").replace(" - vacio", "")
        else:
            cliente = nombre_archivo.replace(".xlsx", "")

        # CORRECCI√ìN: Limpiar cualquier sufijo "- vacio" del nombre del cliente para el t√≠tulo
        cliente_titulo = cliente.replace(" - vacio", "").strip()
        
        print(f"Procesando cliente: {cliente}")

        df = pd.read_excel(excel_file)

        es_archivo_de_html = nombre_archivo.startswith("temp_excel_") or "temp_excel_" in excel_file
        
        if not es_archivo_de_html:            
            impuestos_incluir = [
                'ganancias sociedades',
                'iva',
                'bp-acciones o participaciones',
                'sicore-impto.a las ganancias',
                'empleador-aportes seg. social',
                'contribuciones seg. social',
                'ret art 79 ley gcias inc a,byc',
                'renatea',
                'ganancias personas fisicas',
                'ganancia minima presunta',
                'seguro de vida colectivo',
                'regimenes de informacion',
                'imp s/deb y cred en cta cte',
                'presentac. dj ret. y/o percep',
                'contrib.vales aliment.l.24700',
                'aportes obra social',
                'aseg.riesgo de trabajo l 24557',
                'contribuciones obra social',
                'contribuciones renatea',
                'derecho exportacion servicios',
                'impto.s/bienes personales',
                'multas infracciones formales',
                'retenciones contrib.seg.social',
                'sicore - retenciones y percepc'
            ]

            if 'Impuesto' in df.columns:
                condicion_impuestos = df['Impuesto'].str.contains('|'.join(impuestos_incluir), case=False, na=False)
                df_filtrado = df[condicion_impuestos].copy()
                print(f"Impuestos buscados: {impuestos_incluir}")
                print(f"Registros encontrados con estos impuestos: {len(df_filtrado)}")
            else:
                df_filtrado = df.copy()

            # Obtener la fecha actual y el a√±o actual
            fecha_actual = pd.Timestamp.now().date()
            a√±o_actual = fecha_actual.year
            fecha_inicio = pd.Timestamp(year=a√±o_actual - 7, month=1, day=1).date()  # 1 de enero de 8 a√±os atr√°s
            
            print(f"Rango de fechas: desde {fecha_inicio} hasta {fecha_actual}")

            # Identificar el nombre correcto de la columna de fecha
            columna_fecha_encontrada = None
            posibles_columnas_fecha = ['FechaVencimiento', 'Fecha de Vencimiento', 'Fecha Vencimiento', 'Vencimiento']
            
            for col in posibles_columnas_fecha:
                if col in df_filtrado.columns:
                    columna_fecha_encontrada = col
                    break
            
            # Si encontramos la columna de fecha, procesarla y filtrar
            if columna_fecha_encontrada:
                # Convertir a datetime con formato espec√≠fico y dayfirst=True para formato dd/mm/yyyy
                df_filtrado['fecha_procesada'] = pd.to_datetime(
                    df_filtrado[columna_fecha_encontrada], 
                    errors='coerce',
                    dayfirst=True,  # Especificar que el d√≠a va primero (formato dd/mm/yyyy)
                    format='%d/%m/%Y'  # Especificar el formato expl√≠citamente
                ).dt.date
                
                # Imprimir informaci√≥n de diagn√≥stico
                print(f"Registros antes de filtrar por fecha: {len(df_filtrado)}")
                
                # Filtrar solo por fecha de vencimiento menor a fecha actual (vencido)
                mascara_fechas_validas = df_filtrado['fecha_procesada'].notna()
                
                # Aplicar filtro solo por fecha
                df_filtrado = df_filtrado[
                    mascara_fechas_validas & 
                    (df_filtrado['fecha_procesada'] >= fecha_inicio) &
                    (df_filtrado['fecha_procesada'] <= fecha_actual)
                ]
                
                print(f"Registros despu√©s de filtrar por fecha: {len(df_filtrado)}")
                
                # Eliminar la columna temporal despu√©s de filtrar
                df_filtrado = df_filtrado.drop(['fecha_procesada'], axis=1)
            else:
                print(f"Advertencia: No se encontr√≥ columna de fecha de vencimiento en {excel_file}")
                print(f"Columnas disponibles: {list(df_filtrado.columns)}")
        else:
            # Para archivos que vienen de exportar_desde_html, usar tal como est√°n
            print("Archivo viene de extracci√≥n HTML, usando datos ya filtrados...")
            df_filtrado = df.copy()

        # Verificar si la tabla est√° vac√≠a
        if df_filtrado.shape[0] == 0:
            if " - vacio" not in output_pdf:
                output_pdf = output_pdf.replace(".pdf", " - vacio.pdf")
            print(f"No se encontraron registros que cumplan con los criterios en {excel_file}")

        columnas_a_eliminar = [
            'Int. punitorios', 'Concepto / Subconcepto', 
            'Int. punitorio', 'Int. Punitorio',           
            'Concepto', 'Subconcepto', 'Establecimiento',
            'Fecha_Procesamiento', 'Fuente'               
        ]

        for columna in columnas_a_eliminar:
            if columna in df_filtrado.columns:
                df_filtrado = df_filtrado.drop(columna, axis=1)
                print(f"Columna '{columna}' eliminada en procesar_excel")

        df_filtrado = verificar_columnas_finales(df_filtrado, cliente)
        # Guardar el DataFrame filtrado en el archivo Excel
        df_filtrado.to_excel(excel_file, index=False)

        # FORMATEAR VALORES MONETARIOS ANTES DE GUARDAR EN EXCEL
        print("Formateando valores monetarios individuales...")

        # Formatear columnas monetarias en el DataFrame
        columnas_monetarias = ['Saldo', 'Int. Resarcitorios']
        for col in columnas_monetarias:
            if col in df_filtrado.columns:
                # Mantener valores num√©ricos para c√°lculos internos, formatear solo para visualizaci√≥n
                print(f"Columna {col} preparada para formateo")

        # Cargar el archivo para aplicar formato con openpyxl
        wb = load_workbook(excel_file)
        ws = wb.active

        # Insertar filas adicionales para una nueva imagen
        ws.insert_rows(1, amount=7)

        # Agregar una imagen encima del encabezado (A1)
        # Obtener el ancho combinado de la tabla
        ultima_columna = ws.max_column
        ultima_letra_columna = get_column_letter(ultima_columna)

        # Insertar la imagen
        img = ExcelImage(imagen)
        # Ajustar el tama√±o de la imagen
        img.width = ws.column_dimensions['A'].width * ultima_columna * 6  # Ajustar al ancho combinado
        img.height = 120  # Altura fija
        # Agregar la imagen a la hoja
        ws.add_image(img, 'A1')

        # Insertar filas adicionales para una nueva imagen
        ws.insert_rows(7, amount=1)

        # Fila donde se agregar√° el texto
        fila_texto = 8

        # Obtener el n√∫mero de columnas ocupadas por la tabla
        ultima_columna = ws.max_column
        ultima_letra_columna = get_column_letter(ultima_columna)

        # Combinar celdas en la fila de separaci√≥n
        ws.merge_cells(f'A{fila_texto}:{ultima_letra_columna}{fila_texto}')

        # Establecer el texto en la celda combinada
        celda_texto = ws[f'A{fila_texto}']
        celda_texto.value = f"Reporte de deudas del SCT - {cliente_titulo}"

        # Aplicar formato centrado y en negrita
        celda_texto.alignment = Alignment(horizontal='center', vertical='center')
        celda_texto.font = Font(bold=True, size=20)

        # Cambiar el color del encabezado a lila
        header_fill = PatternFill(start_color="AA0EAA", end_color="AA0EAA", fill_type="solid")
        for cell in ws[9]:
            cell.fill = header_fill

        # Ajustar el ancho de las columnas con control espec√≠fico para "Impuesto"
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            column_header = ""
            
            # Obtener el nombre del encabezado de la columna
            for cell in col:
                if cell.row == 9 and cell.value:  # Fila 9 es donde est√°n los encabezados
                    column_header = str(cell.value).lower()
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Ajuste especial para la columna "Impuesto"
            if "impuesto" in column_header:
                # Limitar el ancho m√°ximo de la columna Impuesto a 35 caracteres
                adjusted_width = min(35, (max_length + 2) * 1.2)
            else:
                # Para el resto de columnas, usar el c√°lculo normal
                adjusted_width = (max_length + 2) * 1.2
            
            ws.column_dimensions[column].width = adjusted_width

        # Encontrar las columnas "Fecha de Vencimiento", "Saldo" e "Int. resarcitorios" para totales y alineaci√≥n
        fecha_vencimiento_col = None
        saldo_col = None
        int_resarcitorios_col = None
        columnas_derecha = []
        header_row = 9  # Fila donde est√°n los encabezados
        
        for col_num, cell in enumerate(ws[header_row], 1):
            if cell.value and isinstance(cell.value, str):
                cell_value_lower = cell.value.lower()
                
                # Buscar columna Fecha de Vencimiento
                if any(term in cell_value_lower for term in ['fecha de vencimiento', 'fechavencimiento', 'vencimiento']):
                    fecha_vencimiento_col = col_num
                
                # Buscar columna Saldo
                elif 'saldo' in cell_value_lower:
                    saldo_col = col_num
                    columnas_derecha.append(col_num)
                
                # Buscar columna Int. resarcitorios
                elif any(term in cell_value_lower for term in ['int. resarcitorios', 'int.resarcitorios', 'resarcitorios']):
                    int_resarcitorios_col = col_num
                    columnas_derecha.append(col_num)
        
        # Aplicar alineaci√≥n a todas las celdas
        for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for col_num, cell in enumerate(row, 1):
                if col_num in columnas_derecha:
                    # Alinear a la derecha las columnas de Saldo e Int. resarcitorios
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    # Centrar el resto de columnas
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Encontrar la √∫ltima fila con datos y agregar fila de totales
        ultima_fila_datos = ws.max_row
        fila_total = ultima_fila_datos + 1
        
        # Agregar "TOTAL" en la columna de fecha de vencimiento
        if fecha_vencimiento_col:
            celda_total = ws.cell(row=fila_total, column=fecha_vencimiento_col)
            celda_total.value = "TOTAL"
            celda_total.font = Font(bold=True)
            celda_total.alignment = Alignment(horizontal='center', vertical='center')
        
        if saldo_col:
            suma_saldo = 0
            print(f"Formateando columna Saldo (columna {saldo_col})")
            
            for fila in range(10, ultima_fila_datos + 1):
                celda_saldo = ws.cell(row=fila, column=saldo_col)
                if celda_saldo.value is not None:
                    try:
                        # Convertir a float para suma
                        valor_numerico = float(celda_saldo.value) if isinstance(celda_saldo.value, (int, float)) else convertir_argentino_a_float(celda_saldo.value)
                        suma_saldo += valor_numerico
                        
                        # Formatear celda individual con formato argentino
                        celda_saldo.value = valor_numerico
                        celda_saldo.number_format = '#,##0.00'  # Excel aplicar√° separadores seg√∫n configuraci√≥n regional
                        
                    except (ValueError, TypeError):
                        print(f"Valor no num√©rico en Saldo fila {fila}: {celda_saldo.value}")
                        celda_saldo.value = 0
                        celda_saldo.number_format = '#,##0.00'
            
            print(f"Total Saldo: {suma_saldo}")
            
            # Insertar la suma con formato
            celda_suma_saldo = ws.cell(row=fila_total, column=saldo_col)
            celda_suma_saldo.value = suma_saldo
            celda_suma_saldo.number_format = '#,##0.00'
            celda_suma_saldo.font = Font(bold=True)
            celda_suma_saldo.alignment = Alignment(horizontal='right', vertical='center')

        # Calcular y agregar sumatoria de Int. resarcitorios con formato
        if int_resarcitorios_col:
            suma_int_resarcitorios = 0
            print(f"Formateando columna Int. Resarcitorios (columna {int_resarcitorios_col})")
            
            for fila in range(10, ultima_fila_datos + 1):
                celda_int = ws.cell(row=fila, column=int_resarcitorios_col)
                if celda_int.value is not None:
                    try:
                        # Convertir a float para suma
                        valor_numerico = float(celda_int.value) if isinstance(celda_int.value, (int, float)) else convertir_argentino_a_float(celda_int.value)
                        suma_int_resarcitorios += valor_numerico
                        
                        # Formatear celda individual
                        celda_int.value = valor_numerico
                        celda_int.number_format = '#,##0.00'
                        
                    except (ValueError, TypeError):
                        print(f"Valor no num√©rico en Int. Resarcitorios fila {fila}: {celda_int.value}")
                        celda_int.value = 0
                        celda_int.number_format = '#,##0.00'
            
            print(f"Total Int. Resarcitorios: {suma_int_resarcitorios}")
            
            # Insertar la suma con formato
            celda_suma_int = ws.cell(row=fila_total, column=int_resarcitorios_col)
            celda_suma_int.value = suma_int_resarcitorios
            celda_suma_int.number_format = '#,##0.00'
            celda_suma_int.font = Font(bold=True)
            celda_suma_int.alignment = Alignment(horizontal='right', vertical='center')

            # AGREGAR FECHA ACTUAL EN FILA 30, COLUMNA INT. RESARCITORIOS
            fecha_actual = datetime.now().strftime("%d/%m/%Y")
            
            celda_fecha = ws.cell(row=38, column=int_resarcitorios_col)
            celda_fecha.value = fecha_actual
            celda_fecha.alignment = Alignment(horizontal='right', vertical='center')
            celda_fecha.font = Font(italic=True)
            print(f"Fecha {fecha_actual} agregada en fila 30, columna Int. Resarcitorios")

        # Guardar los cambios
        wb.save(excel_file)
        ajustar_diseno_excel(ws)
        wb.save(excel_file)
        # Convertir el archivo Excel a PDF con pywin32f
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(excel_file)

        print("Configurando √°rea de impresi√≥n...")
        ws = wb.Worksheets(1)

        # Definir el rango del √°rea de impresi√≥n manualmente
        last_row = ws.UsedRange.Rows.Count
        last_col = ws.UsedRange.Columns.Count
        ws.PageSetup.PrintArea = f"A1:{get_column_letter(last_col)}{last_row + 8}"  # Incluir imagen y tabla

        # Ajustar a una p√°gina
        ws.PageSetup.Orientation = 2  # Paisaje
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = 1

        # Configurar centrado en la p√°gina
        ws.PageSetup.CenterHorizontally = True
        ws.PageSetup.CenterVertically = False

        # Configurar m√°rgenes
        ws.PageSetup.LeftMargin = 0.25
        ws.PageSetup.RightMargin = 0.25
        ws.PageSetup.TopMargin = 0.5
        ws.PageSetup.BottomMargin = 0.5

        print("Guardando como PDF...")
        wb.ExportAsFixedFormat(0, output_pdf)  # 0 indica formato PDF
        wb.Close(False)
        print(f"Archivo convertido a PDF: {output_pdf}")

    except Exception as e:
        print(f"Error al procesar {excel_file}: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if 'excel' in locals():
            excel.Quit()

def iniciar_sesion(cuit_ingresar, password, row_index):
    """Inicia sesi√≥n en el sitio web con el CUIT y contrase√±a proporcionados."""
    try:
        driver.get('https://auth.afip.gob.ar/contribuyente_/login.xhtml')
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'F1:username')))
        element.clear()
        time.sleep(2)

        human_typing(element, cuit_ingresar)
        driver.find_element(By.ID, 'F1:btnSiguiente').click()
        time.sleep(2)

        # Verificar si el CUIT es incorrecto
        try:
            error_message = driver.find_element(By.ID, 'F1:msg').text
            if error_message == "N√∫mero de CUIL/CUIT incorrecto":
                actualizar_excel(row_index, "N√∫mero de CUIL/CUIT incorrecto")
                return False
        except:
            pass

        element_pass = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'F1:password')))
        human_typing(element_pass, password)
        time.sleep(3)
        driver.find_element(By.ID, 'F1:btnIngresar').click()
        time.sleep(2)

        # Verificar si la contrase√±a es incorrecta
        try:
            error_message = driver.find_element(By.ID, 'F1:msg').text
            if error_message == "Clave o usuario incorrecto":
                actualizar_excel(row_index, "Clave incorrecta")
                return False
        except:
            pass

        return True
    except Exception as e:
        print(f"Error al iniciar sesi√≥n: {e}")
        actualizar_excel(row_index, "Error al iniciar sesi√≥n")
        return False

def ingresar_modulo(cuit_ingresar, password, row_index):
    """Ingresa al m√≥dulo espec√≠fico del sistema de cuentas tributarias."""

    # Verificar si el bot√≥n "Ver todos" est√° presente y hacer clic
    boton_ver_todos = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, "Ver todos")))
    if boton_ver_todos:
        boton_ver_todos.click()
        time.sleep(2)

    # Buscar input del buscador y escribir
    buscador = driver.find_element(By.ID, 'buscadorInput')
    if buscador:
        human_typing(buscador, 'tas tr') 
        time.sleep(2)

    # Seleccionar la opci√≥n del men√∫
    opcion_menu = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'rbt-menu-item-0')))
    if opcion_menu:
        opcion_menu.click()
        time.sleep(2)

    # Manejar modal si aparece
    modales = driver.find_elements(By.CLASS_NAME, 'modal-content')
    if modales and modales[0].is_displayed():
        boton_continuar = driver.find_element(By.XPATH, '//button[text()="Continuar"]')
        if boton_continuar:
            boton_continuar.click()
            time.sleep(2)

    # Cambiar a la √∫ltima pesta√±a abierta
    driver.switch_to.window(driver.window_handles[-1])

    # Verificar mensaje de error de autenticaci√≥n
    error_message_elements = driver.find_elements(By.TAG_NAME, 'pre')
    if error_message_elements and error_message_elements[0].text == "Ha ocurrido un error al autenticar, intente nuevamente.":
        actualizar_excel(row_index, "Error autenticacion")
        driver.refresh()
        time.sleep(2)

    # Verificar si es necesario iniciar sesi√≥n nuevamente
    username_input = driver.find_elements(By.ID, 'F1:username')
    if username_input:
        username_input[0].clear()
        time.sleep(2)
        human_typing(username_input[0], cuit_ingresar)
        driver.find_element(By.ID, 'F1:btnSiguiente').click()
        time.sleep(2)

        password_input = driver.find_elements(By.ID, 'F1:password')
        if password_input:
            human_typing(password_input[0], password)
            time.sleep(2)
            driver.find_element(By.ID, 'F1:btnIngresar').click()
            time.sleep(2)
            actualizar_excel(row_index, "Error volver a iniciar sesion")

def seleccionar_cuit_representado(cuit_representado):
    """Selecciona el CUIT representado en el sistema."""
    try:
        select_present = EC.presence_of_element_located((By.NAME, "$PropertySelection"))
        if WebDriverWait(driver, 5).until(select_present):
            current_selection = Select(driver.find_element(By.NAME, "$PropertySelection")).first_selected_option.text
            if current_selection != str(cuit_representado):
                select_element = Select(driver.find_element(By.NAME, "$PropertySelection"))
                select_element.select_by_visible_text(str(cuit_representado))
    except Exception:
        try:
            cuit_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'span.cuit')))
            cuit_text = cuit_element.text.replace('-', '')
            if cuit_text != str(cuit_representado):
                print(f"El CUIT ingresado no coincide con el CUIT representado: {cuit_representado}")
                return False
        except Exception as e:
            print(f"Error al verificar CUIT: {e}")
            return False
    # Esperar que el popup est√© visible y hacer clic en el bot√≥n de cerrar por XPATH
    try:
    # Usamos el XPATH para localizar el bot√≥n de cerrar
        xpath_popup = "/html/body/div[2]/div[2]/div/div/a"
        element_popup = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_popup)))
        element_popup.click()
        print("Popup cerrado exitosamente.")
    except Exception as e:
        print(f"Error al intentar cerrar el popup: {e}")
    return True

def configurar_select_100_mejorado(driver):
    print(f"\n--- CONFIGURANDO SELECT A 100 REGISTROS (VERSI√ìN MEJORADA) ---")
    
    try:
        # Esperar inicial
        time.sleep(2)
        print("‚úì Esperando 2 segundos antes de configurar select...")
        
        # ESTRATEGIA 1: Buscar el select con m√∫ltiples selectores
        select_element = None
        selectores_select = [
            "select.mx-2.form-control.form-control-sm",
            "select[class*='form-control-sm']",
            "select[class*='mx-2']",
            "//div[@class='dtable__footer']//select",
            "//div[contains(@class, 'pagination')]//select",
            "//select[contains(@class, 'form-control')]",
            "//select"  # √öltimo recurso
        ]
        
        for i, selector in enumerate(selectores_select):
            try:
                if selector.startswith("//"):
                    elements = driver.find_elements(By.XPATH, selector)
                else:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                
                if elements:
                    # Verificar cu√°l es el select correcto (que est√© visible y tenga opciones)
                    for element in elements:
                        if element.is_displayed():
                            select_element = element
                            print(f"‚úì Select encontrado con selector {i+1}: {selector}")
                            break
                    
                    if select_element:
                        break
                        
            except Exception as e:
                continue
        
        if not select_element:
            print("‚úó No se encontr√≥ ning√∫n select, continuando sin cambio...")
            time.sleep(2)
            return False
        
        # ESTRATEGIA 2: Analizar el select encontrado
        print(f"\n--- ANALIZANDO SELECT ENCONTRADO ---")
        
        # Hacer scroll al elemento
        driver.execute_script("arguments[0].scrollIntoView(true);", select_element)
        time.sleep(2)
        
        # Obtener informaci√≥n del select
        current_value = select_element.get_attribute('value')
        print(f"Valor actual del select: {current_value}")
        
        # ESTRATEGIA 3: Obtener opciones de manera m√°s robusta
        opciones_info = driver.execute_script("""
            var select = arguments[0];
            var opciones = [];
            
            for (var i = 0; i < select.options.length; i++) {
                var option = select.options[i];
                opciones.push({
                    value: option.value,
                    text: option.text,
                    index: i
                });
            }
            
            return opciones;
        """, select_element)
        
        print(f"Opciones encontradas: {len(opciones_info)}")
        for opcion in opciones_info:
            print(f"  - Valor: '{opcion['value']}', Texto: '{opcion['text']}', √çndice: {opcion['index']}")
        
        # Verificar si ya est√° en 100
        if current_value == "100":
            print("‚úì Select ya est√° configurado en 100")
            time.sleep(2)
            return True
        
        # ESTRATEGIA 4: Buscar la opci√≥n 100
        opcion_100_encontrada = None
        for opcion in opciones_info:
            if opcion['value'] == '100' or opcion['text'] == '100':
                opcion_100_encontrada = opcion
                break
        
        if not opcion_100_encontrada:
            print("‚ö† No se encontr√≥ opci√≥n '100' en el select")
            # Intentar con la opci√≥n m√°s alta disponible
            valores_numericos = []
            for opcion in opciones_info:
                try:
                    if opcion['value'] and opcion['value'].isdigit():
                        valores_numericos.append(int(opcion['value']))
                except:
                    pass
            
            if valores_numericos:
                max_valor = max(valores_numericos)
                print(f"Usando valor m√°ximo disponible: {max_valor}")
                target_value = str(max_valor)
                target_index = None
                for opcion in opciones_info:
                    if opcion['value'] == target_value:
                        target_index = opcion['index']
                        break
            else:
                print("‚úó No se encontraron opciones v√°lidas")
                time.sleep(2)
                return False
        else:
            target_value = "100"
            target_index = opcion_100_encontrada['index']
            print(f"‚úì Opci√≥n 100 encontrada en √≠ndice {target_index}")
        
        # ESTRATEGIA 5: M√∫ltiples m√©todos de cambio
        exito_cambio = False
                   
        # M√©todo 2: Select by index
        if not exito_cambio:
            try:
                print("Intentando M√©todo 2: Select by index...")
                from selenium.webdriver.support.ui import Select
                select_obj = Select(select_element)
                select_obj.select_by_index(target_index)
                time.sleep(2)
                
                new_value = select_element.get_attribute('value')
                if new_value == target_value:
                    print(f"‚úì M√©todo 2 exitoso: Select cambiado a {target_value}")
                    exito_cambio = True
                else:
                    print(f"‚úó M√©todo 2 fall√≥: Valor sigue siendo {new_value}")
                    
            except Exception as e:
                print(f"‚úó M√©todo 2 fall√≥: {e}")
                         
        # ESTRATEGIA 6: Verificaci√≥n visual y de DOM
        if exito_cambio:
            print(f"\n--- VERIFICANDO CAMBIO ---")
            time.sleep(2)
            
            # Verificar valor del select
            valor_final = select_element.get_attribute('value')
            print(f"Valor final del select: {valor_final}")
            
            # Verificar informaci√≥n de paginaci√≥n
            try:
                # Buscar elementos que muestren informaci√≥n de registros
                info_elements = driver.find_elements(By.XPATH, "//*[contains(text(), 'registros') or contains(text(), 'Mostrando') or contains(text(), 'de')]")
                
                for elem in info_elements:
                    if elem.is_displayed():
                        texto = elem.text.strip()
                        if texto and ('registros' in texto.lower() or 'mostrando' in texto.lower()):
                            print(f"Informaci√≥n de paginaci√≥n: {texto}")
                            break
                            
            except Exception as e:
                print(f"No se pudo obtener informaci√≥n de paginaci√≥n: {e}")
            
            # Verificar n√∫mero de filas visibles en la tabla
            try:
                filas_visibles = driver.find_elements(By.XPATH, "//tbody//tr[@role='row']")
                print(f"Filas visibles en la tabla: {len(filas_visibles)}")
                
                if len(filas_visibles) > 10:
                    print("‚úì El cambio parece haber funcionado (m√°s de 10 filas visibles)")
                else:
                    print("‚ö† Posible problema: solo se ven 10 o menos filas")
                    
            except Exception as e:
                print(f"No se pudo contar filas visibles: {e}")
        
        # Esperar antes de continuar
        print("‚úì Esperando 2 segundos antes de extraer datos...")
        time.sleep(2)
        
        return exito_cambio
        
    except Exception as e:
        print(f"‚úó Error general configurando select: {e}")
        time.sleep(2)
        return False

def verificar_columnas_finales(df, cliente):
    # Verifica que solo est√©n las columnas correctas antes de generar PDF.
    print(f"\n--- VERIFICANDO COLUMNAS FINALES PARA {cliente} ---")
    
    columnas_esperadas = ['Impuesto', 'Per√≠odo', 'Ant/Cuota', 'Vencimiento', 'Saldo', 'Int. Resarcitorios']
    columnas_actuales = list(df.columns)
    
    print(f"Columnas actuales: {columnas_actuales}")
    print(f"Columnas esperadas: {columnas_esperadas}")
    
    # Verificar si hay columnas no deseadas
    columnas_extra = [col for col in columnas_actuales if col not in columnas_esperadas]
    if columnas_extra:
        print(f"‚ö† Columnas extra encontradas: {columnas_extra}")
        
        # Eliminar columnas extra
        df_limpio = df[columnas_esperadas].copy()
        print(f"‚úì Columnas extra eliminadas")
        return df_limpio
    else:
        print(f"‚úì Solo columnas correctas presentes")
        return df

def exportar_desde_html(ubicacion_descarga, cuit_representado, cliente):
    try:
        print(f"=== INICIANDO EXTRACCI√ìN HTML PARA CLIENTE: {cliente} ===")
        
        # Verificar que estamos en la p√°gina correcta
        print(f"URL actual: {driver.current_url}")
        print(f"T√≠tulo de la p√°gina: {driver.title}")
        
        # Esperar a que la p√°gina se cargue completamente
        time.sleep(2)
        # PASO 1: Verificar si hay iframe y cambiar a √©l
        print(f"\n--- VERIFICANDO Y CAMBIANDO AL IFRAME ---")
        
        iframe_encontrado = False

        try:
            # Buscar iframe espec√≠fico del SCT
            iframe_selector = "iframe[src*='homeContribuyente']"
            iframe_element = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, iframe_selector)))
            
            print(f"‚úì Iframe encontrado: {iframe_element.get_attribute('src')}")
            
            # Cambiar al iframe
            driver.switch_to.frame(iframe_element)
            iframe_encontrado = True
            print("‚úì Cambiado al iframe exitosamente")
            
            # Esperar a que el contenido del iframe se cargue COMPLETAMENTE
            time.sleep(3)  # Aumentar tiempo de espera
            
            # Esperar a que Vue.js termine de renderizar
            WebDriverWait(driver, 20).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            print("‚úì Contenido del iframe cargado completamente")
            
        except Exception as e:
            print(f"‚úó Error cambiando al iframe: {e}")
            print("Continuando en el documento principal...")
        
        # PASO 2: B√öSQUEDA MEJORADA del elemento "$ Deudas"
        print(f"\n--- B√öSQUEDA MEJORADA DE ELEMENTO '$ DEUDAS' ---")
        
        elemento_deudas = None
        numero_deudas = 0

        try:
            # PRIMERA B√öSQUEDA: Esperar expl√≠citamente a que aparezcan las pesta√±as
            print("Esperando a que las pesta√±as se carguen...")
            
            try:
                # Esperar a que aparezca cualquier elemento de navegaci√≥n
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "[role='tablist'], .nav-tabs, .tab-content"))
                )
                print("‚úì Elementos de navegaci√≥n detectados")
            except:
                print("‚ö† No se detectaron elementos de navegaci√≥n est√°ndar")
            
            # SEGUNDA B√öSQUEDA: Buscar TODOS los elementos que contengan "Deudas"
            print("Buscando TODOS los elementos con 'Deudas'...")
            
            # Usar JavaScript para buscar elementos
            elementos_deudas_js = driver.execute_script("""
                var elementos = [];
                var allElements = document.querySelectorAll('*');
                
                for (var i = 0; i < allElements.length; i++) {
                    var element = allElements[i];
                    if (element.textContent && element.textContent.includes('Deudas')) {
                        elementos.push({
                            tagName: element.tagName,
                            className: element.className,
                            id: element.id,
                            textContent: element.textContent.substring(0, 100),
                            isVisible: element.offsetParent !== null,
                            role: element.getAttribute('role'),
                            href: element.href || ''
                        });
                    }
                }
                
                return elementos;
            """)
            
            print(f"JavaScript encontr√≥ {len(elementos_deudas_js)} elementos con 'Deudas':")
            for i, elem in enumerate(elementos_deudas_js[:10]):  # Mostrar primeros 10
                print(f"  {i+1}. Tag: {elem['tagName']}, Texto: '{elem['textContent'][:50]}...', Visible: {elem['isVisible']}")
                print(f"      Clase: {elem['className']}, Role: {elem['role']}")
            
            # TERCERA B√öSQUEDA: Intentar selectores m√°s amplios
            print("\nBuscando con selectores amplios...")
            
            selectores_amplios = [
                # Buscar cualquier elemento que contenga "Deudas"
                "//*[contains(text(), 'Deudas')]",
                "//*[contains(., 'Deudas')]",
                # Buscar elementos clickeables
                "//a[contains(text(), 'Deudas')]",
                "//button[contains(text(), 'Deudas')]",
                "//div[contains(text(), 'Deudas')]",
                "//span[contains(text(), 'Deudas')]",
                "//li[contains(text(), 'Deudas')]",
                # Buscar por atributos comunes de Bootstrap/Vue
                "//*[@data-*][contains(text(), 'Deudas')]",
                "//*[@v-*][contains(text(), 'Deudas')]",
                # Buscar por clases de Bootstrap
                "//*[contains(@class, 'nav')][contains(text(), 'Deudas')]",
                "//*[contains(@class, 'tab')][contains(text(), 'Deudas')]",
                "//*[contains(@class, 'btn')][contains(text(), 'Deudas')]"
            ]
            
            for i, selector in enumerate(selectores_amplios, 1):
                try:
                    elementos = driver.find_elements(By.XPATH, selector)
                    if elementos:
                        print(f"  Selector {i} encontr√≥ {len(elementos)} elementos")
                        
                        for j, elem in enumerate(elementos):
                            try:
                                if elem.is_displayed():
                                    elem_texto = elem.text.strip()
                                    if 'Deudas' in elem_texto:
                                        print(f"    ‚úì Elemento visible: '{elem_texto}'")
                                        
                                        # Este es nuestro candidato
                                        elemento_deudas = elem
                                        
                                        # Buscar n√∫mero de deudas
                                        import re
                                        numeros = re.findall(r'\d+', elem_texto)
                                        if numeros:
                                            numero_deudas = int(numeros[0])
                                            print(f"    ‚òÖ N√∫mero de deudas: {numero_deudas}")
                                        else:
                                            numero_deudas = 1
                                            
                                        break
                                        
                            except Exception as e:
                                continue
                        
                        if elemento_deudas:
                            break
                            
                except Exception as e:
                    continue
            
            # CUARTA B√öSQUEDA: Si todav√≠a no encuentra, hacer una b√∫squeda exhaustiva
            if not elemento_deudas:
                print("\n--- B√öSQUEDA EXHAUSTIVA ---")
                
                # Guardar HTML completo del iframe para an√°lisis
                iframe_html = driver.page_source
                html_iframe_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"debug_iframe_completo_{cliente}.html")
                with open(html_iframe_file, 'w', encoding='utf-8') as f:
                    f.write(iframe_html)
                print(f"HTML completo del iframe guardado: {html_iframe_file}")
                
                # Buscar "Deudas" en el HTML
                if 'Deudas' in iframe_html:
                    print("‚úì 'Deudas' encontrado en el HTML del iframe")
                    
                    # Intentar hacer clic por coordenadas si es necesario
                    try:
                        # Buscar cualquier elemento que contenga el texto
                        elemento_cualquiera = driver.find_element(By.XPATH, "//*[contains(text(), 'Deudas')]")
                        if elemento_cualquiera:
                            elemento_deudas = elemento_cualquiera
                            numero_deudas = 1
                            print("‚úì Elemento encontrado con b√∫squeda de emergencia")
                    except:
                        pass
                else:
                    print("‚úó 'Deudas' NO encontrado en el HTML del iframe")
                    
        except Exception as e:
            print(f"Error en b√∫squeda de elemento Deudas: {e}")

        if not elemento_deudas:
            print("‚úó No se encontr√≥ el elemento '$ Deudas'")
            
            # Generar PDF vac√≠o y salir
            nombre_pdf = f"Reporte - {cliente} - sin_deudas.pdf"
            ruta_pdf = os.path.join(ubicacion_descarga, nombre_pdf)
            
            df_vacio = pd.DataFrame()
            generar_pdf_desde_dataframe(df_vacio, cliente, ruta_pdf)
            
            # Volver al contenido principal antes de salir
            if iframe_encontrado:
                driver.switch_to.default_content()
            
            return
        
        print(f"‚úì Elemento '$ Deudas' encontrado con {numero_deudas} deudas")

        # PASO 3: Decidir si hacer clic o generar PDF vac√≠o
        datos_tabla = []
        
        if numero_deudas >= 1:
            print(f"\n--- HACIENDO CLIC EN '$ DEUDAS' (tiene {numero_deudas} deudas) ---")
            
            try:
                # Hacer scroll al elemento para asegurar que est√© visible
                driver.execute_script("arguments[0].scrollIntoView(true);", elemento_deudas)
                time.sleep(2)
                
                # Intentar clic normal primero
                elemento_deudas.click()
                print("‚úì Clic normal en '$ Deudas' realizado")
                time.sleep(3)  # Esperar m√°s tiempo para que cargue la tabla

                # USAR LA FUNCI√ìN MEJORADA PARA CONFIGURAR SELECT
                exito_select = configurar_select_100_mejorado(driver)
            
                if not exito_select:
                    print("‚ö† No se pudo configurar el select, continuando con los registros disponibles...")             
            except Exception as e:
                print(f"Error en clic normal: {e}")
                try:
                    # Intentar clic con JavaScript
                    driver.execute_script("arguments[0].click();", elemento_deudas)
                    print("‚úì Clic con JavaScript realizado")
                    time.sleep(2)
                except Exception as e2:
                    print(f"Error en clic JavaScript: {e2}")
                    
                    # Volver al contenido principal antes de salir
                    if iframe_encontrado:
                        driver.switch_to.default_content()
                    return
            
            # PASO 3.5: CONFIGURAR SELECT A 100 REGISTROS
            print(f"\n--- CONFIGURANDO SELECT A 100 REGISTROS ---")

            try:
                time.sleep(2)
                print("‚úì Esperando 2 segundos antes de configurar select...")
                
                # Esperar a que el select est√© presente
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "select.mx-2.form-control.form-control-sm"))
                )
                
                # Buscar el select en el footer de la tabla
                try:
                    select_element = driver.find_element(By.CSS_SELECTOR, "select.mx-2.form-control.form-control-sm")
                    print("‚úì Select encontrado con CSS selector")
                except:
                    # Fallback: buscar por m√∫ltiples selectores
                    selectores_fallback = [
                        "//select[contains(@class, 'form-control-sm')]",
                        "//select[contains(@class, 'mx-2')]", 
                        "//div[@class='dtable__footer']//select",
                        "//div[contains(@class, 'dtable')]//select"
                    ]
                    
                    select_element = None
                    for selector in selectores_fallback:
                        try:
                            select_element = driver.find_element(By.XPATH, selector)
                            print(f"‚úì Select encontrado con selector fallback: {selector}")
                            break
                        except:
                            continue
                    
                    if not select_element:
                        print("‚ö† No se encontr√≥ el select, continuando sin cambiar...")
                        # Continuar sin el select, pero esperar antes de extraer datos
                        time.sleep(2)
                        print("‚úì Esperando 2 segundos antes de extraer datos...")
                    else:
                        # Procesar el select encontrado
                        pass
                
                if 'select_element' in locals() and select_element:
                    # Verificar el valor actual del select
                    current_value = select_element.get_attribute('value')
                    print(f"Valor actual del select: {current_value}")
                    
                    # Buscar todas las opciones disponibles
                    options = select_element.find_elements(By.TAG_NAME, "option")
                    print(f"Opciones disponibles: {[opt.text for opt in options]}")
                    
                    # Verificar si ya est√° en 100
                    if current_value == "100":
                        print("‚úì Select ya est√° configurado en 100")
                    else:
                        # Cambiar a 100
                        try:
                            # M√©todo 1: Usar Select de Selenium
                            from selenium.webdriver.support.ui import Select
                            select_obj = Select(select_element)
                            select_obj.select_by_value("100")
                            print("‚úì Select cambiado a 100 usando Select()")
                            
                        except Exception as e1:
                            print(f"M√©todo 1 fall√≥: {e1}")
                            try:
                                # M√©todo 2: Hacer clic en la opci√≥n 100
                                option_100 = select_element.find_element(By.XPATH, ".//option[@value='100']")
                                option_100.click()
                                print("‚úì Select cambiado a 100 haciendo clic en option")
                                
                            except Exception as e2:
                                print(f"M√©todo 2 fall√≥: {e2}")
                                try:
                                    # M√©todo 3: JavaScript
                                    driver.execute_script("arguments[0].value = '100'; arguments[0].dispatchEvent(new Event('change'));", select_element)
                                    print("‚úì Select cambiado a 100 usando JavaScript")
                                    
                                except Exception as e3:
                                    print(f"M√©todo 3 fall√≥: {e3}")
                                    print("‚ö† No se pudo cambiar el select, continuando...")
                    
                    # Esperar a que la tabla se actualice despu√©s del cambio
                    time.sleep(3)
                    print("‚úì Esperando 3 segundos para que la tabla se actualice...")
                    
                    # Verificar el cambio
                    try:
                        new_value = select_element.get_attribute('value')
                        print(f"Nuevo valor del select: {new_value}")
                        
                        # Buscar el texto que indica cu√°ntos registros se muestran
                        try:
                            registro_text_elements = driver.find_elements(By.XPATH, "//*[contains(text(), 'registros') or contains(text(), 'de')]")
                            for elem in registro_text_elements:
                                if 'registros' in elem.text or 'de' in elem.text:
                                    print(f"Informaci√≥n de registros: {elem.text}")
                                    break
                        except:
                            pass
                            
                    except Exception as e:
                        print(f"Error verificando el cambio: {e}")
                
                # Esperar 2 segundos antes de empezar a extraer datos
                time.sleep(2)
                print("‚úì Esperando 2 segundos antes de extraer datos de la tabla...")

            except Exception as e:
                print(f"Error configurando select: {e}")
                # En caso de error, al menos esperar antes de continuar
                time.sleep(2)
                print("‚úì Esperando 2 segundos antes de continuar (por error en select)...")
            
            # PASO 4: Extraer datos de la tabla (dentro del iframe) - VERSI√ìN OPTIMIZADA
            print(f"\n--- EXTRAYENDO DATOS CON FILTROS COMPLETOS ---")

            try:
                # Esperar a que la tabla se cargue dentro del iframe
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, "//table[@role='table']")))
                
                # Buscar la tabla espec√≠fica con 12 columnas
                tabla = None

                try:
                    tabla = driver.find_element(By.XPATH, "//table[@role='table'][@aria-colcount='12']")
                    aria_rowcount = tabla.get_attribute('aria-rowcount')
                    aria_colcount = tabla.get_attribute('aria-colcount')
                    print(f"‚úì Tabla de 12 columnas encontrada: {aria_rowcount} filas, {aria_colcount} columnas")
                except:
                    # Fallback a b√∫squeda general
                    tablas = driver.find_elements(By.XPATH, "//table[@role='table']")
                    if tablas:
                        tabla = tablas[0]
                        print(f"‚Ñπ Usando primera tabla como fallback")
                    else:
                        print("‚úó No se encontr√≥ tabla")
                        if iframe_encontrado:
                            driver.switch_to.default_content()
                        return
                
                # MAPEO COMPLETO DE TODAS LAS COLUMNAS
                mapeo_columnas_completo = {
                    '1': 'Establecimiento',        # Para luego eliminar
                    '2': 'Concepto',              # Para luego eliminar  
                    '3': 'Subconcepto',           # Para luego eliminar
                    '4': 'Impuesto',              # ‚úì MANTENER
                    '5': 'Concepto',              # Para luego eliminar (duplicado)
                    '6': 'Subconcepto',           # Para luego eliminar (duplicado)  
                    '7': 'Per√≠odo',               # ‚úì MANTENER
                    '8': 'Ant/Cuota',             # ‚úì MANTENER
                    '9': 'Vencimiento',           # ‚úì MANTENER
                    '10': 'Saldo',                # ‚úì MANTENER
                    '11': 'Int. Resarcitorios',   # ‚úì MANTENER
                    '12': 'Int. Punitorio'        # Para luego eliminar
                }
             
                print(f"Mapeo completo definido: {len(mapeo_columnas_completo)} columnas")

                # FILTROS DE IMPUESTOS (igual que en procesar_excel)
                impuestos_incluir = [
                    'ganancias sociedades',
                    'iva',
                    'bp-acciones o participaciones', 
                    'sicore-impto.a las ganancias',
                    'empleador-aportes seg. social',
                    'contribuciones seg. social',
                    'ret art 79 ley gcias inc a,byc',
                    'renatea',
                    'ganancias personas fisicas',
                    'ganancia minima presunta',
                    'seguro de vida colectivo',
                    'regimenes de informacion',
                    'imp s/deb y cred en cta cte',
                    'presentac. dj ret. y/o percep',
                    'contrib.vales aliment.l.24700',
                    'aportes obra social',
                    'aseg.riesgo de trabajo l 24557',
                    'contribuciones obra social',
                    'contribuciones renatea',
                    'derecho exportacion servicios',
                    'impto.s/bienes personales',
                    'multas infracciones formales',
                    'retenciones contrib.seg.social',
                    'sicore - retenciones y percepc'
                ]
                
                print(f"Filtros de impuestos: {impuestos_incluir}")
                
                # CONFIGURAR FECHAS PARA FILTRO
                from datetime import datetime, timedelta
                fecha_actual = datetime.now().date()
                a√±o_actual = fecha_actual.year
                fecha_inicio = datetime(year=a√±o_actual - 7, month=1, day=1).date()
                
                print(f"Filtro de fechas: desde {fecha_inicio} hasta {fecha_actual}")

                # EXTRAER FILAS DE DATOS CON FILTROS
                try:
                    filas_datos = tabla.find_elements(By.XPATH, ".//tbody//tr[@role='row']")
                    print(f"Filas de datos encontradas: {len(filas_datos)}")
                    
                    datos_extraidos = 0
                    datos_filtrados = 0

                    for i, fila in enumerate(filas_datos):
                        try:
                            print(f"\n--- Procesando fila {i+1} ---")
                            
                            # Extraer datos de TODAS las columnas primero
                            datos_fila_completa = {}
                            fila_valida = True
                            for aria_colindex, nombre_columna in mapeo_columnas_completo.items():
                                try:
                                    celda = fila.find_element(By.XPATH, f".//td[@aria-colindex='{aria_colindex}'][@role='cell']")
                                    texto_celda = celda.text.strip()

                                    # Limpiar valores monetarios
                                    if nombre_columna in ['Saldo', 'Int. Resarcitorios', 'Int. Punitorio']:
                                        if not texto_celda or texto_celda in ['', '-', 'N/A']:
                                            texto_celda = 0.0  # Mantener como float para c√°lculos
                                        else:
                                            # Limpiar formato monetario: $ 178.468,79 ‚Üí 178468.79
                                            texto_limpio = texto_celda.replace('$', '').replace(' ', '').strip()
                                            
                                            # Si tiene formato argentino (puntos como separadores de miles, coma como decimal)
                                            if ',' in texto_limpio and '.' in texto_limpio:
                                                # Formato: 178.468,79 ‚Üí 178468.79
                                                partes = texto_limpio.split(',')
                                                if len(partes) == 2:
                                                    parte_entera = partes[0].replace('.', '')
                                                    parte_decimal = partes[1]
                                                    texto_celda = float(f"{parte_entera}.{parte_decimal}")
                                                else:
                                                    texto_celda = float(texto_limpio.replace('.', '').replace(',', '.'))
                                            elif ',' in texto_limpio:
                                                # Solo coma decimal: 1234,56 ‚Üí 1234.56
                                                texto_celda = float(texto_limpio.replace(',', '.'))
                                            elif '.' in texto_limpio:
                                                # Verificar si es separador de miles o decimal
                                                if len(texto_limpio.split('.')[-1]) <= 2:
                                                    # Probablemente decimal
                                                    texto_celda = float(texto_limpio)
                                                else:
                                                    # Probablemente separador de miles
                                                    texto_celda = float(texto_limpio.replace('.', ''))
                                            else:
                                                texto_celda = float(texto_limpio) if texto_limpio else 0.0
                                            
                                            # Validar que sea num√©rico
                                            try:
                                                float(texto_celda)
                                            except (ValueError, TypeError):
                                                texto_celda = 0.0

                                    datos_fila_completa[nombre_columna] = texto_celda
                                    print(f"  {nombre_columna} (col-{aria_colindex}): '{texto_celda}'")
                                except Exception as e:
                                    # Manejo de errores por columna
                                    if nombre_columna in ['Saldo', 'Int. Resarcitorios', 'Int. Punitorio']:
                                        datos_fila_completa[nombre_columna] = '0'
                                        print(f"  {nombre_columna} (col-{aria_colindex}): '0' (por defecto)")
                                    else:
                                        datos_fila_completa[nombre_columna] = ''
                                        print(f"  {nombre_columna} (col-{aria_colindex}): '' (error: {str(e)[:50]}...)")
                                        if nombre_columna in ['Impuesto', 'Vencimiento']:  # Campos cr√≠ticos
                                            fila_valida = False
                            # APLICAR FILTROS
                            if fila_valida:
                                
                                # FILTRO 1: Verificar impuesto
                                impuesto_texto = datos_fila_completa.get('Impuesto', '').lower()
                                impuesto_valido = any(imp in impuesto_texto for imp in impuestos_incluir)
                                
                                if not impuesto_valido:
                                    print(f"  ‚úó Fila {i+1} descartada: impuesto no incluido ('{impuesto_texto}')")
                                    continue
                                
                                # FILTRO 2: Verificar fecha de vencimiento
                                fecha_vencimiento_texto = datos_fila_completa.get('Vencimiento', '')
                                fecha_vencida = False
                                
                                if fecha_vencimiento_texto:
                                    try:
                                        # Parsear fecha formato dd/mm/yyyy
                                        fecha_vencimiento = datetime.strptime(fecha_vencimiento_texto, "%d/%m/%Y").date()
                                        
                                        # Solo incluir si est√° vencida y dentro del rango
                                        if fecha_inicio <= fecha_vencimiento <= fecha_actual:
                                            fecha_vencida = True
                                            print(f"  ‚úì Fecha vencida v√°lida: {fecha_vencimiento}")
                                        else:
                                            print(f"  ‚úó Fecha fuera de rango: {fecha_vencimiento}")
                                            continue
                                            
                                    except ValueError:
                                        print(f"  ‚úó Formato de fecha inv√°lido: '{fecha_vencimiento_texto}'")
                                        continue
                                else:
                                    print(f"  ‚úó Sin fecha de vencimiento")
                                    continue
                                
                                # FILTRO 3: Verificar datos m√≠nimos
                                tiene_datos_minimos = bool(impuesto_texto) and bool(fecha_vencimiento_texto)
                                
                                if tiene_datos_minimos and impuesto_valido and fecha_vencida:
                                    # Agregar metadata de procesamiento
                                    datos_fila_completa['Fecha_Procesamiento'] = fecha_actual.strftime("%Y-%m-%d")
                                    datos_fila_completa['Fuente'] = 'SCT_Web'
                                    
                                    datos_tabla.append(datos_fila_completa)
                                    datos_extraidos += 1
                                    
                                    print(f"  ‚úì Fila {i+1} INCLUIDA en reporte")
                                    print(f"    Resumen: {datos_fila_completa['Impuesto'][:30]}... | {datos_fila_completa['Per√≠odo']} | {datos_fila_completa['Vencimiento']} | ${datos_fila_completa['Saldo']}")
                                else:
                                    print(f"  ‚úó Fila {i+1} descartada: datos insuficientes")
                                    
                            else:
                                print(f"  ‚úó Fila {i+1} descartada: fila inv√°lida")
                            
                            datos_filtrados += 1    
                        except Exception as e:
                            print(f"  ‚úó Error procesando fila {i+1}: {e}")
                            continue

                    print(f"\n‚úì RESUMEN DE EXTRACCI√ìN Y FILTRADO:")
                    print(f"  - Filas procesadas: {len(filas_datos)}")
                    print(f"  - Filas filtradas: {datos_filtrados}")
                    print(f"  - Registros incluidos en reporte: {datos_extraidos}")
                    print(f"  - Tasa de inclusi√≥n: {(datos_extraidos/len(filas_datos)*100):.1f}%" if len(filas_datos) > 0 else "  - Sin filas para procesar")
                    
                    # Mostrar resumen por tipo de impuesto
                    if datos_tabla:
                        impuestos_encontrados = {}
                        for fila in datos_tabla:
                            impuesto = fila['Impuesto']
                            if impuesto in impuestos_encontrados:
                                impuestos_encontrados[impuesto] += 1
                            else:
                                impuestos_encontrados[impuesto] = 1
                        
                        print(f"\n  - Distribuci√≥n por impuesto:")
                        for impuesto, cantidad in impuestos_encontrados.items():
                            print(f"    {impuesto}: {cantidad} registros")
                    
                    # Diagn√≥stico si no se extrajeron datos
                    if datos_extraidos == 0:
                        print(f"\n--- DIAGN√ìSTICO: SIN DATOS EXTRA√çDOS ---")
                        
                        # Verificar una fila de muestra para diagn√≥stico
                        if len(filas_datos) > 0:
                            print("Analizando primera fila para diagn√≥stico...")
                            fila_muestra = filas_datos[0]
                            
                            for aria_colindex, nombre_columna in mapeo_columnas_completo.items():
                                try:
                                    celda = fila_muestra.find_element(By.XPATH, f".//td[@aria-colindex='{aria_colindex}'][@role='cell']")
                                    texto = celda.text.strip()
                                    print(f"    {nombre_columna} (col-{aria_colindex}): '{texto[:50]}...'")
                                except:
                                    print(f"    {nombre_columna} (col-{aria_colindex}): ERROR - No encontrada")
                            
                            # Guardar HTML para an√°lisis
                            tabla_html = tabla.get_attribute('outerHTML')
                            archivo_debug = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"debug_extraccion_filtrada_{cliente}.html")
                            with open(archivo_debug, 'w', encoding='utf-8') as f:
                                f.write(tabla_html)
                            print(f"    HTML guardado para an√°lisis: {archivo_debug}") 

                except Exception as e:
                    print(f"Error extrayendo filas con filtros: {e}")
                    import traceback
                    traceback.print_exc()

            except Exception as e:
                print(f"Error general en extracci√≥n filtrada: {e}")
                import traceback
                traceback.print_exc()
                
                if iframe_encontrado:
                    driver.switch_to.default_content()
                return
        
        # PASO 5: Volver al contenido principal antes de generar PDF
        if iframe_encontrado:
            print("\n--- VOLVIENDO AL CONTENIDO PRINCIPAL ---")
            driver.switch_to.default_content()
            print("‚úì Vuelto al contenido principal")
        # PASO 6: Generar PDF
        print(f"\n--- GENERANDO PDF ---")
        
        nombre_pdf = f"Reporte - {cliente}"
        if not datos_tabla:  # Si no hay datos
            nombre_pdf += " - vacio"
        nombre_pdf += ".pdf"
        
        ruta_pdf = os.path.join(ubicacion_descarga, nombre_pdf)
        
        if datos_tabla:
            df = pd.DataFrame(datos_tabla)
            print(f"DataFrame creado con {len(df)} filas y {len(df.columns)} columnas")
            print(f"Columnas: {list(df.columns)}")
            
            # NO aplicar filtros adicionales - ya est√°n aplicados
            # Los datos ya vienen filtrados por impuesto, fecha y validados
            df_filtrado = df.copy()
            
            print(f"DataFrame final: {len(df_filtrado)} registros para PDF")
            
        else:
            df_filtrado = pd.DataFrame()

        
        # Generar PDF usando la funci√≥n existente (adaptada)
        generar_pdf_desde_dataframe(df_filtrado, cliente, ruta_pdf)
        
        print(f"‚úì PDF generado: {ruta_pdf}")

    except Exception as e:
        print(f"‚úó ERROR GENERAL: {e}")
        import traceback
        traceback.print_exc()
        
        # Asegurar que volvemos al contenido principal en caso de error
        try:
            driver.switch_to.default_content()
        except:
            pass

# MODIFICACI√ìN 2: Nueva funci√≥n unificada para procesar cada cliente
def procesar_cliente_completo(cuit_ingresar, cuit_representado, password, cliente, indice):
    print(f"\n{'='*80}")
    print(f"üöÄ INICIANDO PROCESAMIENTO DE CLIENTE: {cliente}")
    print(f"üìã CUIT Login: {cuit_ingresar} | CUIT Representado: {cuit_representado}")
    print(f"{'='*80}")
    
    try:
        # PASO 1: Configurar navegador nuevo y limpio
        print("üåê PASO 1: Configurando navegador nuevo...")
        configurar_nuevo_navegador()
        
        # PASO 2: Iniciar sesi√≥n
        print("üîê PASO 2: Iniciando sesi√≥n en AFIP...")
        control_sesion = iniciar_sesion(cuit_ingresar, password, indice)
        
        if not control_sesion:
            print(f"‚ùå Error en autenticaci√≥n para {cliente}")
            return False
        
        # PASO 3: Ingresar al m√≥dulo SCT
        print("üè¢ PASO 3: Ingresando al m√≥dulo de Sistema de Cuentas Tributarias...")
        ingresar_modulo(cuit_ingresar, password, indice)
        
        # PASO 4: Cerrar popup inicial
        try:
            xpath_popup = "/html/body/div[2]/div[2]/div/div/a"
            element_popup = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_popup)))
            element_popup.click()
            print("‚úÖ Popup inicial cerrado exitosamente")
        except Exception as e:
            print(f"‚ö†Ô∏è Error al intentar cerrar popup inicial: {e}")
        
        # PASO 5: Seleccionar CUIT representado
        print("üéØ PASO 5: Seleccionando CUIT representado...")
        if not seleccionar_cuit_representado(cuit_representado):
            print(f"‚ùå Error seleccionando CUIT representado para {cliente}")
            return False
        
        # PASO 6: Extraer datos y generar PDF
        print("üìä PASO 6: Extrayendo datos y generando PDF...")
        exportar_desde_html(output_folder_pdf, cuit_representado, cliente)
        
        print(f"‚úÖ CLIENTE {cliente} PROCESADO EXITOSAMENTE")
        return True
        
    except Exception as e:
        print(f"‚ùå ERROR GENERAL procesando cliente {cliente}: {e}")
        import traceback
        traceback.print_exc()
        actualizar_excel(indice, f"Error general: {str(e)[:50]}...")
        return False
    
    finally:
        # PASO 7: SIEMPRE cerrar sesi√≥n y navegador al final
        print("üîí PASO 7: Cerrando sesi√≥n y navegador...")
        cerrar_sesion_y_navegador()
        print(f"üèÅ PROCESAMIENTO DE {cliente} FINALIZADO\n")

# Funci√≥n para convertir Excel a CSV utilizando xlwings
def excel_a_csv(input_folder, output_folder):
    for excel_file in glob.glob(os.path.join(input_folder, "*.xlsx")):
        try:
            app = xw.App(visible=False)
            wb = app.books.open(excel_file)
            sheet = wb.sheets[0]
            df = sheet.used_range.options(pd.DataFrame, header=1, index=False).value

            # Convertir la columna 'FechaVencimiento' a datetime, ajustar seg√∫n sea necesario
            if 'FechaVencimiento' in df.columns:
                df['FechaVencimiento'] = pd.to_datetime(df['FechaVencimiento'], errors='coerce')

            wb.close()
            app.quit()

            base = os.path.basename(excel_file)
            csv_file = os.path.join(output_folder, base.replace('.xlsx', '.csv'))
            df.to_csv(csv_file, index=False, encoding='utf-8-sig', sep=';')
            print(f"Convertido {excel_file} a {csv_file}")
        except Exception as e:
            print(f"Error al convertir {excel_file} a CSV: {e}")

# Funci√≥n para obtener el nombre del cliente a partir del nombre del archivo
def obtener_nombre_cliente(filename):
    base = os.path.basename(filename)
    nombre_cliente = base.split('-')[1].strip()
    return nombre_cliente

def forzar_guardado_excel(excel_file):
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(excel_file)
        wb.Save()
        wb.Close(False)
    except Exception as e:
        print(f"Error forzando guardado en {excel_file}: {e}")
    finally:
        excel.Quit()

def ajustar_diseno_excel(ws):
    # Configurar ajuste de p√°gina para que quepa todo en una p√°gina
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = "landscape"  # Apaisado
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

# ========== VERIFICAR FUNCIONES AL INICIO ==========
print("=" * 60)
print("INICIANDO SISTEMA DE EXTRACCI√ìN DE DEUDAS SCT")
print("=" * 60)
verificar_funciones_disponibles()
print("=" * 60)

# MODIFICACI√ìN 3: Bucle principal simplificado - TODOS los clientes usan sesi√≥n nueva
print("üöÄ INICIANDO PROCESAMIENTO DE CLIENTES")
print("üìã MODO: Sesi√≥n limpia por cliente (sin reutilizaci√≥n)")

indice = 0
for cuit_ingresar, cuit_representado, password, cliente in zip(cuit_login_list, cuit_represent_list, password_list, clientes_list):
    print(f"\nüîÑ PROCESANDO CLIENTE {indice + 1}/{len(clientes_list)}")
    
    # TODOS los clientes ahora usan la funci√≥n unificada
    # No importa si son "nuevos" o "existentes" - siempre sesi√≥n limpia
    exito = procesar_cliente_completo(cuit_ingresar, cuit_representado, password, cliente, indice)
    
    if exito:
        print(f"‚úÖ Cliente {cliente} completado exitosamente")
    else:
        print(f"‚ùå Cliente {cliente} fall√≥ - ver logs para detalles")
    
    indice += 1

print("\n" + "="*60)
print("‚úÖ PROCESAMIENTO DE TODOS LOS CLIENTES COMPLETADO")
print("="*60)

# MODIFICACI√ìN 4: Mantener procesamiento de archivos Excel locales (sin cambios)
print("\nüìÇ PROCESANDO ARCHIVOS EXCEL LOCALES...")

# Recorrer todos los archivos Excel en la carpeta (esto se mantiene para procesar archivos Excel existentes)
for excel_file in glob.glob(os.path.join(input_folder_excel, "*.xlsx")):
    try:
        # Forzar guardado para evitar problemas con archivos corruptos o no calculados
        forzar_guardado_excel(excel_file)

        # Obtener el nombre base del archivo para usarlo en el nombre del PDF
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        output_pdf = os.path.join(output_folder_pdf, f"{base_name}.pdf")
        
        # Llamar a la funci√≥n para procesar el archivo Excel y generar el PDF
        procesar_excel(excel_file, output_pdf, imagen)
        
        print(f"Archivo {excel_file} procesado y guardado como {output_pdf}")
    
    except Exception as e:
        print(f"Error al procesar {excel_file}: {e}")

print("=" * 60)
print("üéâ PROCESO COMPLETADO - ARCHIVOS EXCEL LOCALES")
print("=" * 60)